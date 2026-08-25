#!/usr/bin/env python3
"""Genera un artifact di audit per il lotto Mobilità v6.

Non modifica il catalogo canonico né il sito pubblico. Scarica fonti ufficiali,
verifica i sette Comuni e produce un report HTML/JSON/CSV di collaudo.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import math
import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
import shapefile
from pyproj import CRS, Transformer
from shapely.geometry import Point, shape
from shapely.prepared import prep

TOWNS = [
    ("Camaiore", "046005"),
    ("Forte dei Marmi", "046013"),
    ("Massarosa", "046018"),
    ("Pietrasanta", "046024"),
    ("Seravezza", "046028"),
    ("Stazzema", "046030"),
    ("Viareggio", "046033"),
]
TOWN_NAMES = [x[0] for x in TOWNS]
TOWN_CODES = {name: code for name, code in TOWNS}
SERVICE_DATE = date(2026, 8, 26)

ISTAT_XLSX = {
    2021: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2021_20_03_2026.xlsx",
    2022: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2022_20_03_2026.xlsx",
    2023: "https://www.istat.it/wp-content/uploads/2026/06/Allegato-statistico-2023_20_03_2026.xlsx",
}
BOUNDARIES_URL = "https://www.istat.it/storage/cartografia/confini_amministrativi/generalizzati/2026/Limiti01012026_g.zip"
GTFS_SOURCES = [
    ("bus", "Autolinee Toscane", "https://regionetoscana.smartregion.toscana.it/mobility/artifacts/gtfs"),
    ("rail", "Trenitalia", "https://dati.toscana.it/dataset/8bb8f8fe-fe7d-41d0-90dc-49f2456180d1/resource/4f85393b-357d-443d-8378-65de4198505f/download/trenitalia.gtfs"),
]

HEADERS = {"User-Agent": "Osservatorio-Versilia-mobilita-audit/1.0 (+https://osservatorioversilia.it)"}


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().casefold()
    text = "".join(c for c in unicodedata.normalize("NFKD", text) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", text)


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None
    text = str(value).strip().replace("\u00a0", "").replace("%", "")
    text = text.replace(".", "").replace(",", ".") if re.fullmatch(r"[-+]?\d{1,3}(?:\.\d{3})*(?:,\d+)?", text) else text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def download(session: requests.Session, url: str, target: Path, timeout: int = 180) -> dict[str, Any]:
    target.parent.mkdir(parents=True, exist_ok=True)
    last_error = None
    for attempt in range(4):
        try:
            with session.get(url, stream=True, timeout=(30, timeout), headers=HEADERS) as response:
                response.raise_for_status()
                h = hashlib.sha256()
                size = 0
                with target.open("wb") as fh:
                    for chunk in response.iter_content(chunk_size=1024 * 1024):
                        if not chunk:
                            continue
                        fh.write(chunk)
                        h.update(chunk)
                        size += len(chunk)
                return {
                    "url": url,
                    "sha256": h.hexdigest(),
                    "bytes": size,
                    "contentType": response.headers.get("content-type", ""),
                    "lastModified": response.headers.get("last-modified", ""),
                }
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)
            time.sleep(2 ** attempt)
    raise RuntimeError(f"Download fallito: {url}: {last_error}")


def header_for_column(ws, row_index: int, col_index: int) -> str:
    pieces: list[str] = []
    start = max(1, row_index - 14)
    for r in range(start, row_index):
        value = ws.cell(r, col_index).value
        if value not in (None, ""):
            t = str(value).strip()
            if t and t not in pieces:
                pieces.append(t)
    return " | ".join(pieces[-6:])


def town_row_matches(ws, town: str, code: str) -> list[dict[str, Any]]:
    town_n = norm(town)
    found: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        normalized = [norm(v) for v in values]
        exact_town_cols = [i + 1 for i, v in enumerate(normalized) if v == town_n]
        code_hit = False
        for v in values:
            if v is None:
                continue
            s = re.sub(r"\D", "", str(v))
            if s and s.zfill(6) == code:
                code_hit = True
                break
        if not exact_town_cols and not code_hit:
            continue
        r_idx = row[0].row
        columns = []
        for c_idx, value in enumerate(values, start=1):
            if value in (None, ""):
                continue
            columns.append({
                "column": c_idx,
                "header": header_for_column(ws, r_idx, c_idx),
                "value": value,
            })
        found.append({"row": r_idx, "townColumns": exact_town_cols, "columns": columns})
    return found


def extract_istat_year(path: Path, year: int) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    result: dict[str, Any] = {}
    diagnostics: list[dict[str, Any]] = []
    for town, code in TOWNS:
        candidates = []
        for ws in wb.worksheets:
            matches = town_row_matches(ws, town, code)
            for match in matches:
                vals = match["columns"]
                coe = insist = resident = None
                coe_header = insist_header = resident_header = ""
                for item in vals:
                    h = norm(item["header"])
                    v = num(item["value"])
                    if v is None:
                        continue
                    if "coesist" in h:
                        coe, coe_header = v, item["header"]
                    if "popolazione insistente" in h or ("popolazione" in h and "insistent" in h):
                        if "indice" not in h:
                            insist, insist_header = v, item["header"]
                    if "popolazione residente" in h or ("popolazione" in h and "resident" in h):
                        if "indice" not in h and "non resident" not in h:
                            resident, resident_header = v, item["header"]
                score = 0
                if coe is not None:
                    score += 12
                if insist is not None and resident is not None:
                    score += 8
                if "comun" in norm(ws.title):
                    score += 2
                if match["townColumns"]:
                    score += 1
                if coe is None and insist is not None and resident not in (None, 0):
                    coe = insist / resident * 100.0
                    coe_header = "derivato: popolazione insistente / residente × 100"
                    score += 4
                if coe is not None and 0 < coe <= 3:
                    coe *= 100.0
                candidates.append({
                    "sheet": ws.title,
                    "row": match["row"],
                    "score": score,
                    "coexistenceIndex": coe,
                    "insistentPopulation": insist,
                    "residentPopulation": resident,
                    "headers": {
                        "coexistence": coe_header,
                        "insistent": insist_header,
                        "resident": resident_header,
                    },
                    "preview": [
                        {"header": x["header"], "value": x["value"]}
                        for x in vals[:18]
                    ],
                })
        candidates.sort(key=lambda x: (x["score"], x["coexistenceIndex"] is not None), reverse=True)
        best = candidates[0] if candidates else None
        if best and best["score"] >= 8 and best["coexistenceIndex"] is not None:
            result[town] = {
                "status": "ok",
                "year": year,
                "value": round(float(best["coexistenceIndex"]), 2),
                "insistentPopulation": best["insistentPopulation"],
                "residentPopulation": best["residentPopulation"],
                "sheet": best["sheet"],
                "row": best["row"],
                "method": best["headers"]["coexistence"],
            }
        else:
            result[town] = {"status": "nd", "year": year, "value": None}
        diagnostics.append({
            "town": town,
            "year": year,
            "selected": best,
            "candidateCount": len(candidates),
            "topCandidates": candidates[:5],
        })
    return result, diagnostics


def find_shapefile(folder: Path) -> Path:
    shp_files = list(folder.rglob("*.shp"))
    preferred = [p for p in shp_files if "com" in p.name.casefold()]
    if preferred:
        return preferred[0]
    if not shp_files:
        raise RuntimeError("Nessuno shapefile trovato nei confini Istat")
    return shp_files[0]


def load_town_polygons(boundary_zip: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    extract_dir = boundary_zip.parent / "boundaries"
    with zipfile.ZipFile(boundary_zip) as zf:
        zf.extractall(extract_dir)
    shp_path = find_shapefile(extract_dir)
    reader = shapefile.Reader(str(shp_path), encoding="utf-8")
    fields = [f[0] for f in reader.fields[1:]]
    name_idx = next((i for i, f in enumerate(fields) if f.casefold() in {"comune", "den_com", "nome_com"}), None)
    code_idx = next((i for i, f in enumerate(fields) if f.casefold() in {"pro_com_t", "pro_com", "cod_com", "codice_com"}), None)
    if code_idx is None:
        code_idx = next((i for i, f in enumerate(fields) if "pro_com" in f.casefold() or "cod_com" in f.casefold()), None)
    if name_idx is None:
        name_idx = next((i for i, f in enumerate(fields) if "comune" in f.casefold()), None)
    if code_idx is None:
        raise RuntimeError(f"Campo codice comunale non trovato. Campi: {fields}")
    prj = shp_path.with_suffix(".prj")
    crs = CRS.from_wkt(prj.read_text(encoding="utf-8", errors="ignore")) if prj.exists() else CRS.from_epsg(32632)
    transformer = Transformer.from_crs(CRS.from_epsg(4326), crs, always_xy=True)
    polygons: dict[str, Any] = {}
    meta: dict[str, Any] = {"shapefile": str(shp_path.name), "fields": fields, "crs": crs.to_string()}
    target_codes = {code: name for name, code in TOWNS}
    for sr in reader.iterShapeRecords():
        raw_code = sr.record[code_idx]
        digits = re.sub(r"\D", "", str(raw_code))
        code = digits.zfill(6) if digits else ""
        if code not in target_codes:
            continue
        name = target_codes[code]
        geom = shape(sr.shape.__geo_interface__)
        polygons[name] = prep(geom)
        if name_idx is not None:
            meta.setdefault("names", {})[name] = str(sr.record[name_idx])
    missing = sorted(set(TOWN_NAMES) - set(polygons))
    if missing:
        raise RuntimeError(f"Confini mancanti per: {missing}")
    return {"prepared": polygons, "transformer": transformer}, meta


def member_name(zf: zipfile.ZipFile, basename: str) -> str | None:
    basename = basename.casefold()
    for name in zf.namelist():
        if Path(name).name.casefold() == basename:
            return name
    return None


def csv_from_zip(zf: zipfile.ZipFile, basename: str):
    member = member_name(zf, basename)
    if member is None:
        return None, None
    raw = zf.open(member, "r")
    text = io.TextIOWrapper(raw, encoding="utf-8-sig", newline="")
    return csv.DictReader(text), text


def parse_gtfs_date(value: str) -> date | None:
    value = (value or "").strip()
    if not re.fullmatch(r"\d{8}", value):
        return None
    return datetime.strptime(value, "%Y%m%d").date()


def active_services(zf: zipfile.ZipFile, service_date: date) -> tuple[set[str], dict[str, Any]]:
    active: set[str] = set()
    coverage_dates: list[date] = []
    weekday = service_date.strftime("%A").casefold()
    weekday_map = {
        "monday": "monday", "tuesday": "tuesday", "wednesday": "wednesday",
        "thursday": "thursday", "friday": "friday", "saturday": "saturday", "sunday": "sunday",
    }
    reader, text = csv_from_zip(zf, "calendar.txt")
    if reader:
        for row in reader:
            start, end = parse_gtfs_date(row.get("start_date", "")), parse_gtfs_date(row.get("end_date", ""))
            if start:
                coverage_dates.append(start)
            if end:
                coverage_dates.append(end)
            if start and end and start <= service_date <= end and row.get(weekday_map[weekday], "0") == "1":
                active.add(row.get("service_id", ""))
        text.close()
    exceptions: dict[str, str] = {}
    reader, text = csv_from_zip(zf, "calendar_dates.txt")
    if reader:
        for row in reader:
            d = parse_gtfs_date(row.get("date", ""))
            if d:
                coverage_dates.append(d)
            if d == service_date:
                exceptions[row.get("service_id", "")] = row.get("exception_type", "")
        text.close()
    for service_id, ex_type in exceptions.items():
        if ex_type == "1":
            active.add(service_id)
        elif ex_type == "2":
            active.discard(service_id)
    coverage = {
        "start": min(coverage_dates).isoformat() if coverage_dates else None,
        "end": max(coverage_dates).isoformat() if coverage_dates else None,
        "dateCovered": bool(coverage_dates and min(coverage_dates) <= service_date <= max(coverage_dates)),
        "activeServiceIds": len(active),
    }
    return active, coverage


def parse_time_seconds(value: str) -> int | None:
    m = re.fullmatch(r"(\d{1,2}):(\d{2}):(?:\d{2})", (value or "").strip())
    if not m:
        return None
    return int(m.group(1)) * 3600 + int(m.group(2)) * 60


def fmt_gtfs_time(seconds: int | None) -> str | None:
    if seconds is None:
        return None
    return f"{seconds // 3600:02d}:{(seconds % 3600) // 60:02d}"


def audit_gtfs_feed(path: Path, mode: str, label: str, polygons_bundle: dict[str, Any]) -> dict[str, Any]:
    prepared = polygons_bundle["prepared"]
    transformer = polygons_bundle["transformer"]
    town_stop_ids: dict[str, set[str]] = {town: set() for town in TOWN_NAMES}
    stop_towns: dict[str, list[str]] = defaultdict(list)
    with zipfile.ZipFile(path) as zf:
        active, coverage = active_services(zf, SERVICE_DATE)
        reader, text = csv_from_zip(zf, "stops.txt")
        if reader is None:
            raise RuntimeError(f"{label}: stops.txt mancante")
        stop_count = 0
        for row in reader:
            stop_count += 1
            lat, lon = num(row.get("stop_lat")), num(row.get("stop_lon"))
            sid = row.get("stop_id", "")
            if lat is None or lon is None or not sid:
                continue
            x, y = transformer.transform(lon, lat)
            p = Point(x, y)
            for town in TOWN_NAMES:
                if prepared[town].covers(p):
                    stop_towns[sid].append(town)
                    town_stop_ids[town].add(sid)
                    break
        text.close()

        reader, text = csv_from_zip(zf, "trips.txt")
        if reader is None:
            raise RuntimeError(f"{label}: trips.txt mancante")
        active_trips: dict[str, str] = {}
        for row in reader:
            if row.get("service_id", "") in active:
                tid = row.get("trip_id", "")
                if tid:
                    active_trips[tid] = row.get("route_id", "")
        text.close()

        trip_sets: dict[str, set[str]] = {town: set() for town in TOWN_NAMES}
        route_sets: dict[str, set[str]] = {town: set() for town in TOWN_NAMES}
        active_stop_sets: dict[str, set[str]] = {town: set() for town in TOWN_NAMES}
        first_time: dict[str, int | None] = {town: None for town in TOWN_NAMES}
        last_time: dict[str, int | None] = {town: None for town in TOWN_NAMES}

        reader, text = csv_from_zip(zf, "stop_times.txt")
        if reader is None:
            raise RuntimeError(f"{label}: stop_times.txt mancante")
        scanned = matched = 0
        for row in reader:
            scanned += 1
            trip_id = row.get("trip_id", "")
            if trip_id not in active_trips:
                continue
            sid = row.get("stop_id", "")
            towns = stop_towns.get(sid)
            if not towns:
                continue
            pickup = (row.get("pickup_type") or "0").strip()
            dropoff = (row.get("drop_off_type") or "0").strip()
            if pickup == "1" and dropoff == "1":
                continue
            matched += 1
            t_sec = parse_time_seconds(row.get("departure_time") or row.get("arrival_time") or "")
            for town in towns:
                trip_sets[town].add(trip_id)
                active_stop_sets[town].add(sid)
                route_id = active_trips.get(trip_id, "")
                if route_id:
                    route_sets[town].add(route_id)
                if t_sec is not None:
                    if first_time[town] is None or t_sec < first_time[town]:
                        first_time[town] = t_sec
                    if last_time[town] is None or t_sec > last_time[town]:
                        last_time[town] = t_sec
        text.close()

    towns = {}
    for town in TOWN_NAMES:
        towns[town] = {
            "mode": mode,
            "source": label,
            "status": "ok" if coverage["dateCovered"] else "nd",
            "allGeocodedStops": len(town_stop_ids[town]),
            "activeAccessPoints": len(active_stop_sets[town]) if coverage["dateCovered"] else None,
            "trips": len(trip_sets[town]) if coverage["dateCovered"] else None,
            "routes": len(route_sets[town]) if coverage["dateCovered"] else None,
            "first": fmt_gtfs_time(first_time[town]) if coverage["dateCovered"] else None,
            "last": fmt_gtfs_time(last_time[town]) if coverage["dateCovered"] else None,
            "firstSeconds": first_time[town],
            "lastSeconds": last_time[town],
        }
    return {
        "mode": mode,
        "label": label,
        "coverage": coverage,
        "stopsScanned": stop_count,
        "activeTrips": len(active_trips),
        "stopTimesScanned": scanned,
        "stopTimesMatchedVersilia": matched,
        "towns": towns,
    }


def recursive_dicts(value: Any):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from recursive_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from recursive_dicts(child)


def extract_residents_from_site_data(path: Path) -> tuple[dict[str, int | None], dict[str, Any]]:
    residents = {town: None for town in TOWN_NAMES}
    diagnostics: list[dict[str, Any]] = []
    if not path.exists():
        return residents, {"status": "missing-file"}
    data = json.loads(path.read_text(encoding="utf-8"))
    candidates = []
    for obj in recursive_dicts(data):
        rows = obj.get("rows")
        if not isinstance(rows, list):
            continue
        meta = obj.get("meta", {}) if isinstance(obj.get("meta"), dict) else {}
        labels = " ".join(str(x) for x in [meta.get("key"), meta.get("label"), meta.get("title"), meta.get("name"), obj.get("label"), obj.get("title")] if x)
        score = 0
        nlabels = norm(labels)
        if "popolazione residente" in nlabels:
            score += 10
        if any(k in nlabels for k in ["population", "popolazione"]):
            score += 3
        row_towns = {norm(r.get("town")): r for r in rows if isinstance(r, dict) and r.get("town")}
        score += len(set(norm(t) for t in TOWN_NAMES) & set(row_towns))
        if score >= 8:
            candidates.append((score, obj, labels))
    candidates.sort(key=lambda x: x[0], reverse=True)
    for score, obj, labels in candidates:
        tmp = {}
        for row in obj.get("rows", []):
            if not isinstance(row, dict):
                continue
            town = next((t for t in TOWN_NAMES if norm(t) == norm(row.get("town"))), None)
            if not town:
                continue
            value = num(row.get("value"))
            if value is not None:
                tmp[town] = int(round(value))
        diagnostics.append({"score": score, "labels": labels, "townsFound": sorted(tmp)})
        if len(tmp) >= 6:
            residents.update(tmp)
            break
    return residents, {"status": "ok" if any(v is not None for v in residents.values()) else "not-found", "candidates": diagnostics[:5]}


def combine_gtfs(feeds: list[dict[str, Any]], residents: dict[str, int | None]) -> dict[str, Any]:
    rows = {}
    for town in TOWN_NAMES:
        modes = {f["mode"]: f["towns"][town] for f in feeds}
        all_ok = all(x["status"] == "ok" for x in modes.values())
        bus = modes.get("bus", {})
        rail = modes.get("rail", {})
        if all_ok:
            bus_trips = int(bus.get("trips") or 0)
            rail_trips = int(rail.get("trips") or 0)
            total = bus_trips + rail_trips
            access = int(bus.get("activeAccessPoints") or 0) + int(rail.get("activeAccessPoints") or 0)
            routes = int(bus.get("routes") or 0) + int(rail.get("routes") or 0)
            first_candidates = [x.get("firstSeconds") for x in modes.values() if x.get("firstSeconds") is not None]
            last_candidates = [x.get("lastSeconds") for x in modes.values() if x.get("lastSeconds") is not None]
            first_s = min(first_candidates) if first_candidates else None
            last_s = max(last_candidates) if last_candidates else None
            pop = residents.get(town)
            per1000 = total / pop * 1000.0 if pop and pop > 0 else None
            span_h = (last_s - first_s) / 3600.0 if first_s is not None and last_s is not None else None
            rows[town] = {
                "status": "ok",
                "trips": total,
                "tripsPer1000": round(per1000, 2) if per1000 is not None else None,
                "busTrips": bus_trips,
                "railTrips": rail_trips,
                "activeAccessPoints": access,
                "routes": routes,
                "first": fmt_gtfs_time(first_s),
                "last": fmt_gtfs_time(last_s),
                "serviceSpanHours": round(span_h, 2) if span_h is not None else None,
                "residentPopulation": pop,
                "zeroIsCertified": total == 0,
            }
        else:
            rows[town] = {
                "status": "nd", "trips": None, "tripsPer1000": None,
                "busTrips": bus.get("trips"), "railTrips": rail.get("trips"),
                "activeAccessPoints": None, "routes": None, "first": None, "last": None,
                "serviceSpanHours": None, "residentPopulation": residents.get(town), "zeroIsCertified": False,
            }
    return rows


def fmt(value: Any, decimals: int = 0) -> str:
    if value is None:
        return "n.d."
    if isinstance(value, float):
        s = f"{value:,.{decimals}f}"
    else:
        s = f"{value:,}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value))


def build_html(audit: dict[str, Any]) -> str:
    istat = audit["istat"]["years"]
    gtfs = audit["gtfs"]["combined"]
    coe23_vals = [v["value"] for v in istat["2023"].values() if v.get("value") is not None]
    max_gtfs = max([r["tripsPer1000"] or 0 for r in gtfs.values()] + [1])
    rows_coe = []
    rows_tpl = []
    bars = []
    for town in TOWN_NAMES:
        y21, y22, y23 = istat["2021"][town], istat["2022"][town], istat["2023"][town]
        rows_coe.append(f"<tr><th>{esc(town)}</th><td>{fmt(y21.get('value'),1)}</td><td>{fmt(y22.get('value'),1)}</td><td><b>{fmt(y23.get('value'),1)}</b></td><td>{esc(y23.get('sheet') or '—')}</td></tr>")
        r = gtfs[town]
        status = '<span class="ok">OK</span>' if r["status"] == "ok" else '<span class="nd">n.d.</span>'
        rows_tpl.append(
            f"<tr><th>{esc(town)}</th><td>{status}</td><td><b>{fmt(r.get('trips'))}</b></td><td>{fmt(r.get('tripsPer1000'),2)}</td>"
            f"<td>{fmt(r.get('busTrips'))}</td><td>{fmt(r.get('railTrips'))}</td><td>{fmt(r.get('activeAccessPoints'))}</td>"
            f"<td>{fmt(r.get('routes'))}</td><td>{esc(r.get('first') or 'n.d.')}</td><td>{esc(r.get('last') or 'n.d.')}</td></tr>"
        )
        width = (r.get("tripsPer1000") or 0) / max_gtfs * 100
        bars.append(f'<div class="bar-row"><span>{esc(town)}</span><div class="bar"><i style="width:{width:.1f}%"></i></div><b>{fmt(r.get("tripsPer1000"),2)}</b></div>')

    coe_coverage = sum(1 for x in istat["2023"].values() if x.get("status") == "ok")
    tpl_coverage = sum(1 for x in gtfs.values() if x.get("status") == "ok")
    coe_decision = "GO" if coe_coverage >= 6 else "HOLD"
    tpl_decision = "GO" if tpl_coverage >= 6 else "HOLD"
    generated = audit["generatedAt"]
    source_cards = "".join(
        f'<div class="source"><b>{esc(k)}</b><small>SHA256 {esc(v.get("sha256", ""))[:16]}… · {fmt(v.get("bytes"))} byte</small></div>'
        for k, v in audit["sources"].items()
    )
    return f"""<!doctype html>
<html lang="it"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Audit Mobilità v6 · Osservatorio Versilia</title>
<style>
:root{{--ink:#18202b;--muted:#687386;--line:#dfe5ec;--bg:#f5f7fa;--card:#fff;--accent:#1864ab;--good:#18794e;--bad:#b42318}}
*{{box-sizing:border-box}}body{{margin:0;background:var(--bg);color:var(--ink);font-family:Inter,system-ui,-apple-system,Segoe UI,sans-serif}}
main{{max-width:1180px;margin:auto;padding:28px 18px 60px}}h1{{font-size:clamp(30px,5vw,52px);margin:.15em 0}}h2{{margin-top:34px}}p{{line-height:1.55;color:var(--muted)}}
.hero,.card,.table-wrap,.source{{background:var(--card);border:1px solid var(--line);border-radius:16px;box-shadow:0 3px 18px #172b4d0a}}.hero{{padding:28px}}.eyebrow{{font-weight:800;letter-spacing:.08em;text-transform:uppercase;color:var(--accent);font-size:12px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:14px;margin:18px 0}}.card{{padding:20px}}.card strong{{font-size:28px;display:block;margin:6px 0}}.pill{{display:inline-block;padding:4px 9px;border-radius:999px;font-weight:800;font-size:12px;background:#e8f5ee;color:var(--good)}}
.table-wrap{{overflow:auto}}table{{border-collapse:collapse;width:100%;min-width:760px}}th,td{{padding:12px 14px;border-bottom:1px solid var(--line);text-align:right;white-space:nowrap}}th:first-child,td:first-child{{text-align:left}}thead th{{font-size:12px;text-transform:uppercase;color:var(--muted);background:#f9fafb}}.ok{{color:var(--good);font-weight:800}}.nd{{color:var(--bad);font-weight:800}}
.bar-row{{display:grid;grid-template-columns:160px 1fr 72px;gap:10px;align-items:center;margin:9px 0}}.bar{{height:12px;background:#e8edf3;border-radius:999px;overflow:hidden}}.bar i{{display:block;height:100%;background:var(--accent);border-radius:999px}}.bar-row b{{text-align:right}}
.note{{border-left:4px solid var(--accent);padding:12px 16px;background:#eef5ff;border-radius:0 10px 10px 0}}.source{{padding:14px}}.source small{{display:block;color:var(--muted);margin-top:4px;word-break:break-all}}code{{background:#eef1f5;padding:2px 5px;border-radius:5px}}footer{{margin-top:40px;color:var(--muted);font-size:13px}}
@media(max-width:650px){{.bar-row{{grid-template-columns:110px 1fr 60px}}}}
</style></head><body><main>
<section class="hero"><div class="eyebrow">Artifact di collaudo · non pubblicato</div><h1>Mobilità · audit candidati v6</h1><p>Generato il {esc(generated)} dall'attuale branch di audit. Questo artifact non modifica i 146 indicatori canonici né il sito pubblico. Obiettivo: verificare dati, copertura 7/7, regole 0/n.d. e forma dei due candidati prima della PR implementativa.</p></section>
<div class="grid">
<div class="card"><span class="pill">{coe_decision}</span><strong>{coe_coverage}/7</strong><b>Indice di coesistenza</b><p>Istat, popolazione insistente per studio e lavoro. Candidato a <b>+1 indicatore</b>.</p></div>
<div class="card"><span class="pill">{tpl_decision}</span><strong>{tpl_coverage}/7</strong><b>Offerta TPL programmata</b><p>GTFS Regione Toscana + Trenitalia, giorno tipo {SERVICE_DATE.strftime('%d/%m/%Y')}. Candidato a <b>+1 indicatore</b>.</p></div>
<div class="card"><span class="pill">AGGREGA</span><strong>0</strong><b>Nuove card accessorie</b><p>Fermate, linee, bus/ferrovia e ampiezza oraria restano dettagli del blocco TPL.</p></div>
<div class="card"><span class="pill">RINVIA</span><strong>2</strong><b>Seconda tranche</b><p>Accessibilità della popolazione al TPL e puntualità realtime richiedono GIS/storico dedicato.</p></div>
</div>
<h2>1. Indice di coesistenza della popolazione insistente</h2>
<p>Valore 100 = popolazione insistente uguale ai residenti; valori superiori indicano maggiore attrazione netta per studio/lavoro. Turismo escluso dalla quantificazione corrente.</p>
<div class="table-wrap"><table><thead><tr><th>Comune</th><th>2021</th><th>2022</th><th>2023</th><th>Foglio selezionato</th></tr></thead><tbody>{''.join(rows_coe)}</tbody></table></div>
<h2>2. Offerta TPL programmata</h2>
<p>Una <code>trip_id</code> attiva conta una sola volta per Comune se serve almeno un punto GTFS nel territorio con salita o discesa consentita. Bus e ferrovia sono sommati nel totale; lo zero è ammesso solo quando la fonte e la data sono valide.</p>
<div class="table-wrap"><table><thead><tr><th>Comune</th><th>Stato</th><th>Corse</th><th>/1.000 res.</th><th>Bus</th><th>Ferrovia</th><th>Punti accesso</th><th>Linee</th><th>Prima</th><th>Ultima</th></tr></thead><tbody>{''.join(rows_tpl)}</tbody></table></div>
<div class="card" style="margin-top:16px"><b>Confronto corse / 1.000 residenti</b><div style="margin-top:14px">{''.join(bars)}</div></div>
<h2>3. Decisione editoriale proposta</h2>
<div class="grid"><div class="card"><b>Nuovi indicatori canonici: 2</b><p>Indice di coesistenza + Offerta TPL programmata. Se entrambi superano il gate, catalogo 146 → 148.</p></div><div class="card"><b>Aggregati nel blocco TPL</b><p>Corse assolute, bus/ferrovia, punti di accesso, linee e ampiezza oraria. Nessuna moltiplicazione artificiale delle card.</p></div><div class="card"><b>Scarto</b><p>Frequenza media comunale: metrica troppo dipendente da pesi e fasce orarie.</p></div></div>
<p class="note"><b>Regola 0 / n.d.</b> — <b>0</b> soltanto se feed, data, geometria e calcolo sono validi e il conteggio è realmente nullo. <b>n.d.</b> se manca o non è validabile uno di questi passaggi.</p>
<h2>4. Fonti e riproducibilità</h2><div class="grid">{source_cards}</div>
<footer>Per il controllo tecnico completo aprire anche <code>audit.json</code>, <code>istat.csv</code>, <code>tpl.csv</code> e <code>diagnostics.json</code> presenti nello stesso artifact.</footer>
</main></body></html>"""


def write_csvs(out: Path, audit: dict[str, Any]) -> None:
    with (out / "istat.csv").open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(["Comune", "2021", "2022", "2023", "status_2023", "sheet_2023", "row_2023"])
        for town in TOWN_NAMES:
            y = audit["istat"]["years"]
            writer.writerow([town, y["2021"][town].get("value"), y["2022"][town].get("value"), y["2023"][town].get("value"), y["2023"][town].get("status"), y["2023"][town].get("sheet"), y["2023"][town].get("row")])
    with (out / "tpl.csv").open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(["Comune", "status", "residenti", "corse", "corse_x_1000", "bus", "ferrovia", "punti_accesso", "linee", "prima", "ultima", "ampiezza_ore", "zero_certificato"])
        for town in TOWN_NAMES:
            r = audit["gtfs"]["combined"][town]
            writer.writerow([town, r.get("status"), r.get("residentPopulation"), r.get("trips"), r.get("tripsPer1000"), r.get("busTrips"), r.get("railTrips"), r.get("activeAccessPoints"), r.get("routes"), r.get("first"), r.get("last"), r.get("serviceSpanHours"), r.get("zeroIsCertified")])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="artifacts/mobilita-v6-review")
    args = parser.parse_args()
    out = Path(args.output)
    if out.exists():
        shutil.rmtree(out)
    out.mkdir(parents=True)
    session = requests.Session()
    sources: dict[str, Any] = {}
    diagnostics: dict[str, Any] = {}

    with tempfile.TemporaryDirectory(prefix="mobilita-v6-") as td:
        temp = Path(td)
        istat_years: dict[str, Any] = {}
        istat_diag: list[dict[str, Any]] = []
        for year, url in ISTAT_XLSX.items():
            target = temp / f"istat-{year}.xlsx"
            sources[f"Istat allegato {year}"] = download(session, url, target)
            values, diag = extract_istat_year(target, year)
            istat_years[str(year)] = values
            istat_diag.extend(diag)

        boundary_zip = temp / "confini-2026.zip"
        sources["Istat confini 2026"] = download(session, BOUNDARIES_URL, boundary_zip)
        polygons, boundary_meta = load_town_polygons(boundary_zip)
        diagnostics["boundaries"] = boundary_meta

        feeds = []
        for mode, label, url in GTFS_SOURCES:
            target = temp / f"{mode}.gtfs"
            sources[f"GTFS {label}"] = download(session, url, target, timeout=600)
            feeds.append(audit_gtfs_feed(target, mode, label, polygons))

        residents, resident_diag = extract_residents_from_site_data(Path("data/site-data.json"))
        diagnostics["residentPopulation"] = resident_diag
        combined = combine_gtfs(feeds, residents)

    diagnostics["istatExtraction"] = istat_diag
    audit = {
        "schemaVersion": 1,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "repositoryBaseline": os.environ.get("GITHUB_SHA", "local"),
        "serviceDate": SERVICE_DATE.isoformat(),
        "towns": [{"name": n, "code": c} for n, c in TOWNS],
        "rules": {
            "zero": "Solo con fonte/data/geometria valide e conteggio realmente nullo.",
            "nd": "Fonte, data o attribuzione non disponibile/non validabile.",
            "coverageStandard": "7/7; 6/7 ammesso solo con un unico dato ufficiale mancante, senza stima.",
        },
        "sources": sources,
        "istat": {"years": istat_years},
        "gtfs": {"feeds": feeds, "combined": combined},
        "residents": residents,
    }
    (out / "audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    (out / "diagnostics.json").write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_csvs(out, audit)
    (out / "index.html").write_text(build_html(audit), encoding="utf-8")
    (out / "README.txt").write_text(
        "AUDIT MOBILITA V6 - NON PUBBLICATO\n\nAprire index.html.\n"
        "audit.json contiene i valori completi e la manifestazione delle fonti.\n"
        "diagnostics.json conserva le evidenze di estrazione, incluse le righe Istat selezionate.\n"
        "istat.csv e tpl.csv sono esportazioni rapide per il controllo.\n",
        encoding="utf-8",
    )
    print(f"Artifact generato in {out}")
    print(json.dumps({
        "istat2023Coverage": sum(1 for x in istat_years["2023"].values() if x.get("status") == "ok"),
        "tplCoverage": sum(1 for x in combined.values() if x.get("status") == "ok"),
        "serviceDate": SERVICE_DATE.isoformat(),
    }, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
