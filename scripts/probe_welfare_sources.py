#!/usr/bin/env python3
"""Probe official Welfare / first-infancy sources for the seven Versilia municipalities.

Diagnostic only. The municipal Welfare gate uses the current Istat "A misura di
Comune" workbooks (10a/10b); the 2022 release workbook is retained as a
methodological cross-check because its published tables are mainly aggregate.
Regione Toscana 2024/25 is checked municipality by municipality.
"""
from __future__ import annotations

import csv
import io
import json
import re
import ssl
import unicodedata
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "artifacts" / "welfare-probe"
OUT.mkdir(parents=True, exist_ok=True)

ISTAT_MUNICIPAL = {
    "10a": "https://www.istat.it/storage/misura-comune/10a-Servizi-sociali-per-tipologia-di-utenza.xlsx",
    "10b": "https://www.istat.it/storage/misura-comune/10b-Servizi-sociali-per-abitante.xlsx",
}
ISTAT_2022_RELEASE = "https://www.istat.it/wp-content/uploads/2025/09/Tavole_2022_spesa_comuni-1.xlsx"
TOSCANA_URL = "https://dati.toscana.it/dataset/98ee6064-b61a-45e2-a790-86c55b278574/resource/01588909-8f0b-4b80-8af7-1749bab80a5e/download/opendata-_-da-pubblicare-24-25.csv"
TOWNS = ["Camaiore", "Forte dei Marmi", "Massarosa", "Pietrasanta", "Seravezza", "Stazzema", "Viareggio"]
KEYWORDS = [
    "comune", "spesa", "abitante", "pro capite", "pro-capite", "totale",
    "famiglie", "minori", "disabil", "anzian", "povert", "immigr",
    "dipenden", "multiutenza", "utente", "utenza", "ricett", "posti",
    "bambini", "popolazione", "servizi educativi",
]


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().lower()


def fetch(url: str, target: Path) -> None:
    req = urllib.request.Request(url, headers={"User-Agent": "OsservatorioVersilia/1.0"})
    with urllib.request.urlopen(req, timeout=180, context=ssl.create_default_context()) as response:
        target.write_bytes(response.read())


def compact_row(values: list[object]) -> list[object]:
    last = max((i for i, value in enumerate(values) if value not in (None, "")), default=-1)
    return values[: last + 1] if last >= 0 else []


def probe_workbook(path: Path, url: str) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {"url": url, "sheets": [], "townCoverage": {town: [] for town in TOWNS}}
    towns_norm = {norm(t): t for t in TOWNS}
    for ws in wb.worksheets:
        sheet = {"title": ws.title, "maxRow": ws.max_row, "maxColumn": ws.max_column, "keywordRows": [], "townRows": []}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = compact_row(list(row))
            if not values:
                continue
            joined = " | ".join(norm(v) for v in values if v not in (None, ""))
            if len(sheet["keywordRows"]) < 40 and any(k in joined for k in KEYWORDS):
                sheet["keywordRows"].append({"row": r_idx, "values": values[:80]})
            matched = []
            for cell in values:
                n = norm(cell)
                for town_n, town in towns_norm.items():
                    if town_n and (n == town_n or town_n in n):
                        matched.append(town)
            if matched:
                item = {"row": r_idx, "towns": sorted(set(matched)), "values": values[:100]}
                sheet["townRows"].append(item)
                for town in set(matched):
                    result["townCoverage"][town].append({"sheet": ws.title, "row": r_idx})
        result["sheets"].append(sheet)
    return result


def decode_csv(raw: bytes) -> tuple[str, str]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            pass
    raise RuntimeError("Impossibile decodificare CSV Toscana")


def probe_toscana(path: Path) -> dict:
    text, encoding = decode_csv(path.read_bytes())
    dialect = csv.Sniffer().sniff(text[:12000], delimiters=";,\t|")
    rows = list(csv.reader(io.StringIO(text), dialect))
    header = rows[0] if rows else []
    matches = []
    coverage = {town: 0 for town in TOWNS}
    town_norm = {norm(t): t for t in TOWNS}
    for idx, row in enumerate(rows[1:], start=2):
        joined = " | ".join(norm(v) for v in row)
        found = [town for town_n, town in town_norm.items() if town_n in joined]
        if found:
            matches.append({"row": idx, "towns": sorted(set(found)), "values": row})
            for town in set(found):
                coverage[town] += 1
    return {"url": TOSCANA_URL, "encoding": encoding, "delimiter": dialect.delimiter, "header": header,
            "rowCount": max(0, len(rows) - 1), "townRows": matches, "townCoverage": coverage}


def report_workbook(lines: list[str], title: str, probe: dict) -> None:
    lines += ["", f"## {title}"]
    for sheet in probe["sheets"]:
        if sheet["townRows"] or sheet["keywordRows"]:
            lines += [f"### {sheet['title']}", f"- dimensione: {sheet['maxRow']} × {sheet['maxColumn']}",
                      f"- righe Versilia trovate: {len(sheet['townRows'])}"]
            for item in sheet["keywordRows"][:12]:
                lines.append(f"- header/keyword riga {item['row']}: `{item['values']}`")
            for item in sheet["townRows"][:30]:
                lines.append(f"- riga {item['row']} · {', '.join(item['towns'])}: `{item['values']}`")
    lines.append("### Copertura")
    for town, hits in probe["townCoverage"].items():
        lines.append(f"- {town}: {len(hits)} occorrenze")


def main() -> None:
    municipal = {}
    for key, url in ISTAT_MUNICIPAL.items():
        path = OUT / f"istat-{key}.xlsx"
        fetch(url, path)
        municipal[key] = probe_workbook(path, url)

    release_path = OUT / "istat-2022-release.xlsx"
    fetch(ISTAT_2022_RELEASE, release_path)
    release = probe_workbook(release_path, ISTAT_2022_RELEASE)

    toscana_path = OUT / "toscana-prima-infanzia-2024-25.csv"
    fetch(TOSCANA_URL, toscana_path)
    toscana = probe_toscana(toscana_path)

    result = {"istatMunicipal": municipal, "istat2022Release": release, "toscana": toscana}
    (OUT / "probe.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")

    lines = ["# Welfare / prima infanzia · probe fonti"]
    report_workbook(lines, "Istat A misura di Comune · 10a tipologia utenza", municipal["10a"])
    report_workbook(lines, "Istat A misura di Comune · 10b spesa per abitante", municipal["10b"])
    lines += ["", "## Istat release 2022 · controllo metodologico",
              "Il file della release 2022 viene ispezionato ma non è usato come gate comunale: le tavole pubblicate sono prevalentemente aggregate."]
    report_workbook(lines, "Istat release 2022 · tavole", release)

    t = toscana
    lines += ["", "## Regione Toscana · Prima infanzia 2024/25", f"- intestazioni: `{t['header']}`", f"- righe: {t['rowCount']}", "### Righe Versilia"]
    for item in t["townRows"]:
        lines.append(f"- riga {item['row']} · {', '.join(item['towns'])}: `{item['values']}`")
    lines.append("### Copertura Toscana")
    for town, count in t["townCoverage"].items():
        lines.append(f"- {town}: {count} righe")
    (OUT / "probe.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    missing = {}
    for key in ("10a", "10b"):
        missing[key] = [town for town, hits in municipal[key]["townCoverage"].items() if not hits]
    missing["toscana"] = [town for town, count in toscana["townCoverage"].items() if count == 0]
    print(f"Istat 10a missing: {missing['10a']}")
    print(f"Istat 10b missing: {missing['10b']}")
    print(f"Toscana missing: {missing['toscana']}")
    if any(missing.values()):
        raise SystemExit(2)


if __name__ == "__main__":
    main()
